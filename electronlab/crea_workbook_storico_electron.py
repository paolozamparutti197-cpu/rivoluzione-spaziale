import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


API_URL = "https://ll.thespacedevs.com/2.3.0/launches/previous/"
OUTPUT = Path(__file__).with_name("lanci_electron.xlsx")
ROME = ZoneInfo("Europe/Rome")

TEAL = "0F4C5C"
NAVY = "0B1F33"
ORANGE = "F59E0B"
PALE = "EAF2F5"
WHITE = "FFFFFF"
GREEN = "D9EAD3"
RED = "F4CCCC"
GRID = "C7D4DA"


def safe_get(data, *keys, default=""):
    current = data
    for key in keys:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
        if current is None:
            return default
    return current


def clean_text(value):
    text = " ".join(str(value or "").split())
    if "â" in text:
        try:
            text = text.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return text


def fetch_launches(retries=3):
    params = {
        "format": "json",
        "mode": "detailed",
        "rocket__configuration__name": "Electron",
        "ordering": "net",
        "limit": 100,
    }
    for attempt in range(retries):
        response = requests.get(API_URL, params=params, timeout=90)
        if response.status_code != 429 or attempt == retries - 1:
            break
        wait_seconds = int(response.headers.get("Retry-After", 20 * (attempt + 1)))
        print(f"Rate limit The Space Devs: attendo {wait_seconds}s...")
        time.sleep(wait_seconds)
    response.raise_for_status()
    data = response.json()
    launches = data.get("results", [])
    if data.get("next"):
        raise RuntimeError("La risposta API richiede paginazione: aumentare o gestire il limite.")
    return launches


def first_url(items):
    for item in items or []:
        if isinstance(item, dict) and item.get("url"):
            return item["url"]
    return ""


def launch_row(number, launch):
    net = datetime.fromisoformat(launch["net"].replace("Z", "+00:00"))
    mission = launch.get("mission") or {}
    pad = launch.get("pad") or {}
    location = pad.get("location") or {}
    status_name = safe_get(launch, "status", "name")
    status = "successo" if status_name == "Launch Successful" else "fallito"
    agencies = ", ".join(
        clean_text(item.get("name")) for item in mission.get("agencies") or [] if item.get("name")
    )
    video = first_url(launch.get("vid_urls")) or first_url(mission.get("vid_urls"))
    mission_name = clean_text(mission.get("name"))
    if not mission_name:
        mission_name = clean_text(str(launch.get("name") or "").split("|")[-1])
    return {
        "nr": number,
        "data": net.date(),
        "ora_utc": net.strftime("%H:%M"),
        "anno": net.year,
        "missione": mission_name,
        "stato": status,
        "tipo_missione": clean_text(mission.get("type")),
        "orbita": clean_text(safe_get(mission, "orbit", "abbrev") or safe_get(mission, "orbit", "name")),
        "cliente": agencies,
        "pad": clean_text(pad.get("name")),
        "localita": clean_text(location.get("name")),
        "descrizione": clean_text(mission.get("description")),
        "causa_fallimento": clean_text(launch.get("failreason")),
        "video": video,
        "fonte": launch.get("url") or "",
        "id_launch_library": launch.get("id") or "",
    }


def title_row(ws, title, subtitle, end_column):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    ws["A1"] = title
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, italic=True, color="40515C")
    ws["A2"].fill = PatternFill("solid", fgColor=PALE)
    ws.row_dimensions[2].height = 23


def style_header(ws, row, start, end):
    thin = Side(style="thin", color=GRID)
    for column in range(start, end + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[row].height = 30


def add_table(ws, ref, name):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def build_workbook(rows):
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "Storico lanci Electron - Rocket Lab"
    wb.properties.subject = "Riepilogo dei lanci Electron"
    wb.properties.creator = "Rivoluzione Spaziale"

    elenco = wb.create_sheet("elenco")
    headers = [
        "nr", "data", "ora UTC", "anno", "missione", "stato", "tipo missione", "orbita",
        "cliente", "pad", "localita", "descrizione", "causa fallimento", "video", "fonte LL2",
        "id Launch Library 2",
    ]
    for column, header in enumerate(headers, 1):
        elenco.cell(1, column, header)
    style_header(elenco, 1, 1, len(headers))
    for row_index, item in enumerate(rows, 2):
        values = list(item.values())
        for column, value in enumerate(values, 1):
            cell = elenco.cell(row_index, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=column in {5, 9, 11, 12, 13})
            if column == 2:
                cell.number_format = "dd/mm/yyyy"
            if column == 6:
                cell.fill = PatternFill("solid", fgColor=GREEN if value == "successo" else RED)
                cell.font = Font(bold=True, color="274E13" if value == "successo" else "990000")
        for column in (14, 15):
            cell = elenco.cell(row_index, column)
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
    elenco.freeze_panes = "A2"
    elenco.auto_filter.ref = f"A1:P{len(rows) + 1}"
    widths = [7, 12, 9, 8, 34, 12, 22, 12, 24, 28, 30, 58, 48, 18, 18, 38]
    for index, width in enumerate(widths, 1):
        elenco.column_dimensions[get_column_letter(index)].width = width
    add_table(elenco, f"A1:P{len(rows) + 1}", "StoricoElectron")

    successes = sum(item["stato"] == "successo" for item in rows)
    failures = len(rows) - successes
    years = Counter(item["anno"] for item in rows)
    pads = Counter(item["pad"] or item["localita"] or "Non indicato" for item in rows)
    orbits = Counter(item["orbita"] or "Non indicata" for item in rows)

    dashboard = wb.create_sheet("dashboard")
    title_row(
        dashboard,
        "ROCKET LAB · STORICO LANCI ELECTRON",
        f"Dati aggiornati al {datetime.now(ROME):%d/%m/%Y %H:%M} · Fonte: The Space Devs / Launch Library 2",
        10,
    )
    metrics = [
        ("Lanci", len(rows)), ("Successi", successes), ("Fallimenti", failures),
        ("Tasso di successo", successes / len(rows) if rows else 0),
        ("Primo lancio", rows[0]["data"] if rows else ""),
        ("Ultimo lancio", rows[-1]["data"] if rows else ""),
        ("Lanci nel 2026", years.get(2026, 0)), ("Pad utilizzati", len(pads)),
    ]
    for index, (label, value) in enumerate(metrics):
        column = 1 + (index % 4) * 2
        row = 4 + (index // 4) * 3
        dashboard.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        dashboard.cell(row, column, label)
        dashboard.cell(row, column).fill = PatternFill("solid", fgColor=TEAL)
        dashboard.cell(row, column).font = Font(bold=True, color=WHITE)
        dashboard.cell(row, column).alignment = Alignment(horizontal="center")
        dashboard.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
        dashboard.cell(row + 1, column, value)
        dashboard.cell(row + 1, column).fill = PatternFill("solid", fgColor=PALE)
        dashboard.cell(row + 1, column).font = Font(size=17, bold=True, color=NAVY)
        dashboard.cell(row + 1, column).alignment = Alignment(horizontal="center")
        if label == "Tasso di successo":
            dashboard.cell(row + 1, column).number_format = "0.0%"
        if "lancio" in label.lower() and isinstance(value, datetime):
            dashboard.cell(row + 1, column).number_format = "dd/mm/yyyy"
        elif label in {"Primo lancio", "Ultimo lancio"}:
            dashboard.cell(row + 1, column).number_format = "dd/mm/yyyy"
    dashboard["A11"] = "Ultima missione"
    dashboard["A11"].font = Font(bold=True, color=WHITE)
    dashboard["A11"].fill = PatternFill("solid", fgColor=ORANGE)
    dashboard.merge_cells("A11:B11")
    dashboard.merge_cells("C11:J11")
    dashboard["C11"] = f"#{rows[-1]['nr']} · {rows[-1]['missione']} · {rows[-1]['data']:%d/%m/%Y}" if rows else ""
    dashboard["C11"].font = Font(bold=True, color=NAVY)
    dashboard["C11"].fill = PatternFill("solid", fgColor="FFF2CC")
    dashboard["A13"] = "Fonte dati"
    dashboard["B13"] = API_URL
    dashboard["B13"].hyperlink = API_URL
    dashboard["B13"].style = "Hyperlink"
    for column in range(1, 11):
        dashboard.column_dimensions[get_column_letter(column)].width = 15
    dashboard.sheet_view.showGridLines = False

    annuale = wb.create_sheet("serie_annuale")
    title_row(annuale, "LANCI ELECTRON PER ANNO", "Successi, fallimenti e tasso di successo", 5)
    annual_headers = ["anno", "lanci", "successi", "fallimenti", "tasso successo"]
    for column, header in enumerate(annual_headers, 1):
        annuale.cell(4, column, header)
    style_header(annuale, 4, 1, 5)
    annual_rows = []
    for year in sorted(years):
        subset = [item for item in rows if item["anno"] == year]
        ok = sum(item["stato"] == "successo" for item in subset)
        annual_rows.append((year, len(subset), ok, len(subset) - ok, ok / len(subset)))
    for row_index, values in enumerate(annual_rows, 5):
        for column, value in enumerate(values, 1):
            annuale.cell(row_index, column, value)
        annuale.cell(row_index, 5).number_format = "0.0%"
    add_table(annuale, f"A4:E{4 + len(annual_rows)}", "SerieAnnualeElectron")
    chart = BarChart()
    chart.title = "Lanci Electron per anno"
    chart.y_axis.title = "Lanci"
    chart.add_data(Reference(annuale, min_col=2, min_row=4, max_row=4 + len(annual_rows)), titles_from_data=True)
    chart.set_categories(Reference(annuale, min_col=1, min_row=5, max_row=4 + len(annual_rows)))
    chart.height = 8
    chart.width = 16
    annuale.add_chart(chart, "G4")
    for column, width in enumerate([12, 12, 12, 14, 18], 1):
        annuale.column_dimensions[get_column_letter(column)].width = width

    per_pad = wb.create_sheet("per_pad")
    title_row(per_pad, "LANCI ELECTRON PER PAD", "Distribuzione dei lanci completati", 4)
    for column, header in enumerate(["pad", "lanci", "quota", "ultimo utilizzo"], 1):
        per_pad.cell(4, column, header)
    style_header(per_pad, 4, 1, 4)
    pad_rows = []
    for pad, count in pads.most_common():
        last_date = max(item["data"] for item in rows if (item["pad"] or item["localita"] or "Non indicato") == pad)
        pad_rows.append((pad, count, count / len(rows), last_date))
    for row_index, values in enumerate(pad_rows, 5):
        for column, value in enumerate(values, 1):
            per_pad.cell(row_index, column, value)
        per_pad.cell(row_index, 3).number_format = "0.0%"
        per_pad.cell(row_index, 4).number_format = "dd/mm/yyyy"
    add_table(per_pad, f"A4:D{4 + len(pad_rows)}", "PadElectron")
    per_pad.column_dimensions["A"].width = 45
    for column in "BCD":
        per_pad.column_dimensions[column].width = 18

    per_orbita = wb.create_sheet("per_orbita")
    title_row(per_orbita, "LANCI ELECTRON PER ORBITA", "Classificazione orbitale indicata da Launch Library 2", 3)
    for column, header in enumerate(["orbita", "lanci", "quota"], 1):
        per_orbita.cell(4, column, header)
    style_header(per_orbita, 4, 1, 3)
    orbit_rows = [(orbit, count, count / len(rows)) for orbit, count in orbits.most_common()]
    for row_index, values in enumerate(orbit_rows, 5):
        for column, value in enumerate(values, 1):
            per_orbita.cell(row_index, column, value)
        per_orbita.cell(row_index, 3).number_format = "0.0%"
    add_table(per_orbita, f"A4:C{4 + len(orbit_rows)}", "OrbiteElectron")
    pie = PieChart()
    pie.title = "Orbite Electron"
    pie.add_data(Reference(per_orbita, min_col=2, min_row=4, max_row=4 + len(orbit_rows)), titles_from_data=True)
    pie.set_categories(Reference(per_orbita, min_col=1, min_row=5, max_row=4 + len(orbit_rows)))
    pie.height = 8
    pie.width = 12
    per_orbita.add_chart(pie, "E4")
    per_orbita.column_dimensions["A"].width = 24
    per_orbita.column_dimensions["B"].width = 14
    per_orbita.column_dimensions["C"].width = 14

    legenda = wb.create_sheet("legenda")
    title_row(legenda, "LEGENDA E FONTI", "Note metodologiche del workbook Electron", 5)
    notes = [
        ("Perimetro", "Tutti i lanci completati con configurazione del lanciatore Electron."),
        ("Fonte primaria", "The Space Devs / Launch Library 2 API."),
        ("Filtro API", "rocket__configuration__name=Electron sull'endpoint launches/previous."),
        ("Data e ora", "Data del T-0 in UTC; nel foglio elenco la data e l'ora UTC sono separate."),
        ("Successo", "Status Launch Successful nella fonte."),
        ("Fallimento", "Status Launch Failure nella fonte."),
        ("Aggiornamento", "Eseguire crea_workbook_storico_electron.py per rigenerare il file."),
    ]
    for row_index, (label, note) in enumerate(notes, 4):
        legenda.cell(row_index, 1, label)
        legenda.cell(row_index, 1).font = Font(bold=True, color=TEAL)
        legenda.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=5)
        legenda.cell(row_index, 2, note)
        legenda.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
    legenda.column_dimensions["A"].width = 22
    for column in "BCDE":
        legenda.column_dimensions[column].width = 24
    legenda["B5"].hyperlink = API_URL
    legenda["B5"].style = "Hyperlink"

    wb.active = wb.sheetnames.index("dashboard")
    return wb


def main():
    launches = fetch_launches()
    rows = [launch_row(index, launch) for index, launch in enumerate(launches, 1)]
    workbook = build_workbook(rows)
    workbook.save(OUTPUT)
    successes = sum(item["stato"] == "successo" for item in rows)
    print(f"Creato: {OUTPUT}")
    print(f"Lanci Electron: {len(rows)} · successi: {successes} · fallimenti: {len(rows) - successes}")
    if rows:
        print(f"Ultimo lancio: #{rows[-1]['nr']} · {rows[-1]['missione']} · {rows[-1]['data']:%d/%m/%Y}")


if __name__ == "__main__":
    main()
