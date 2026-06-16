import json
import subprocess
import ast
from datetime import date, datetime, timezone
from html import escape
from pathlib import Path
from urllib.parse import quote

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SECTIONS_DIR = ROOT / "sezioni"
CSS_DIR = ROOT / "css"

HERO_IMAGE = "https://upload.wikimedia.org/wikipedia/commons/5/5d/Falcon_1_Flight_4_launch.jpg"
CSS_VERSION = "20260614-nav-gerarchia"

MAIN_SECTIONS = [
    {
        "title": "SpaceX",
        "slug": "spacex",
        "status": "Attiva",
        "copy": "Cadenza Falcon, riuso, agenda lanci e sviluppo Starship.",
        "active": True,
    },
    {
        "title": "Blue Origin",
        "slug": "blue-origin",
        "status": "In costruzione",
        "copy": "New Glenn, BE-4 e il tentativo di costruire un secondo polo privato pesante.",
    },
    {
        "title": "ULA",
        "slug": "ula",
        "status": "In costruzione",
        "copy": "Vulcan, sicurezza nazionale e transizione dalla stagione Atlas/Delta.",
    },
    {
        "title": "Rocket Lab",
        "slug": "rocket-lab",
        "status": "In costruzione",
        "copy": "Electron, Neutron e la via agile al lancio commerciale integrato.",
    },
    {
        "title": "Arianespace",
        "slug": "arianespace",
        "status": "In costruzione",
        "copy": "Ariane 6, Vega C e la ricerca europea di autonomia d'accesso.",
    },
    {
        "title": "Cina",
        "slug": "cina",
        "status": "In costruzione",
        "copy": "Lunga Marcia, stazione spaziale, Luna e nuovo ecosistema commerciale.",
    },
]

THEME_SECTIONS = [
    {
        "title": "Luna",
        "slug": "luna",
        "status": "In costruzione",
        "copy": "Artemis, lander commerciali, basi, risorse e competizione geopolitica.",
    },
    {
        "title": "Marte",
        "slug": "marte",
        "status": "In costruzione",
        "copy": "Dalla retorica della colonizzazione alle tecnologie davvero necessarie.",
    },
    {
        "title": "Infrastrutture orbitali",
        "slug": "infrastrutture-orbitali",
        "status": "In costruzione",
        "copy": "Stazioni private, rifornimento, cargo, servicing e logistica in orbita.",
    },
    {
        "title": "Cronologia",
        "slug": "cronologia",
        "status": "In costruzione",
        "copy": "Una linea del tempo per seguire la rivoluzione spaziale senza perdere il filo.",
    },
]

UTILITY_SECTIONS = [
    {
        "title": "Lanci imminenti",
        "slug": "lanci-imminenti",
        "status": "Attiva",
        "copy": "Agenda SpaceX con countdown e fonti missione.",
    },
    {
        "title": "Storico lanci",
        "slug": "storico-lanci",
        "status": "Attiva",
        "copy": "Dashboard Falcon: cadenza, famiglie, pad e recupero booster.",
    },
    {
        "title": "Starship",
        "slug": "starship",
        "status": "Attiva",
        "copy": "Sviluppo Starship, voli integrati e temi tecnici principali.",
    },
    {
        "title": "Pad di lancio",
        "slug": "pad-di-lancio",
        "status": "Attiva",
        "copy": "Mappa dei complessi SpaceX, schede operative e infografiche dei pad.",
    },
    {
        "title": "Localita SpaceX",
        "slug": "localita-spacex",
        "status": "Attiva",
        "copy": "Mappa delle principali localita SpaceX negli Stati Uniti, inclusa la proposta Louisiana.",
    },
]

PLACEHOLDER_SECTIONS = MAIN_SECTIONS + THEME_SECTIONS


def clean(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    if isinstance(value, float):
        return round(value, 4)
    return value


def rows_from_sheet(path, sheet_name, header_row):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[header_row]]
    rows = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None for value in row):
            continue
        item = {}
        for index, header in enumerate(headers):
            if header is None or index >= len(row):
                continue
            item[str(header)] = clean(row[index])
        rows.append(item)
    return rows


def extract_upcoming_launches():
    node_code = r"""
const fs = require("fs");
const text = fs.readFileSync("spacex_lanci_fino_luglio_2026 (1).html", "utf8");
const match = text.match(/const launches = \[([\s\S]*?)\];/);
if (!match) throw new Error("launches not found");
const launches = Function("return [" + match[1] + "];")();
console.log(JSON.stringify(launches.map((launch) => ({
  id: launch.id,
  cat: launch.cat,
  name: launch.name,
  status: launch.status,
  rocket: launch.rocket,
  dateLabel: launch.dateLabel,
  iso: launch.iso,
  window: launch.window,
  site: launch.site,
  landing: launch.landing,
  orbit: launch.orbit,
  payload: launch.payload,
  summary: launch.summary,
  news: launch.news,
  sources: (launch.sources || []).map((source) => ({ label: source[0], url: source[1] }))
}))));
"""
    try:
        output = subprocess.check_output(["node", "-e", node_code], cwd=ROOT, text=True, encoding="utf-8")
        launches = json.loads(output)
    except Exception:
        return []
    now = datetime.now(timezone.utc)
    upcoming = []
    for launch in launches:
        categories = launch.get("cat") or []
        iso = launch.get("iso")
        if "exact" in categories and iso:
            try:
                launch_time = datetime.fromisoformat(str(iso).replace("Z", "+00:00"))
            except ValueError:
                launch_time = None
            if launch_time and launch_time <= now:
                continue
        upcoming.append(launch)
    return sorted(upcoming, key=lambda item: str(item.get("iso") or "9999"))


def falcon_data():
    path = ROOT / "01_workbook" / "lanci_spacex_falcon.xlsx"
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    dashboard = workbook["dashboard"]
    elenco = workbook["elenco"]
    metrics = {
        "lanci": clean(dashboard["A4"].value),
        "successi": clean(dashboard["C4"].value),
        "tasso": clean(dashboard["E4"].value),
        "boosterMax": clean(dashboard["G4"].value),
        "falconHeavy": clean(dashboard["I4"].value),
        "padAttivi": clean(dashboard["K4"].value),
        "primo": clean(dashboard["A7"].value),
        "ultimo": clean(dashboard["C7"].value),
        "recordAggiuntivi": clean(dashboard["E7"].value),
        "tentativiRecupero": clean(dashboard["G7"].value),
        "recuperiRiusciti": clean(dashboard["I7"].value),
        "boosterRiutilizzati": clean(dashboard["K7"].value),
    }
    annual = [
        {
            "anno": row.get("anno"),
            "lanci": row.get("lanci principali"),
            "successi": row.get("successi"),
            "falliti": row.get("falliti o parziali"),
            "tasso": row.get("tasso successo"),
        }
        for row in rows_from_sheet(path, "serie_annuale", 3)
        if row.get("anno") is not None
    ]
    launchers = [
        {
            "lanciatore": row.get("lanciatore"),
            "lanci": row.get("lanci principali"),
            "successi": row.get("successi"),
            "tasso": row.get("tasso successo"),
        }
        for row in rows_from_sheet(path, "per_lanciatore", 3)
        if row.get("lanciatore")
    ][:7]
    pad_sheet = workbook["per_pad_orbita"]
    pads = [
        {"pad": clean(row[0]), "lanci": clean(row[1]), "quota": row[2]}
        for row in pad_sheet.iter_rows(min_row=4, max_col=3, values_only=True)
        if row[0]
    ]
    landing_rows = rows_from_sheet(path, "booster_landing", 3)
    landing = [
        {
            "codice": row.get("codice landing"),
            "record": row.get("record"),
            "tentativi": row.get("tentativi"),
            "recuperi": row.get("recuperi riusciti"),
        }
        for row in landing_rows
        if row.get("codice landing")
    ][:8]
    latest = {}
    for values in elenco.iter_rows(min_row=2, values_only=True):
        data = values[1] if len(values) > 1 else None
        lanciatore = values[3] if len(values) > 3 else None
        cliente = values[4] if len(values) > 4 else None
        if data and lanciatore and cliente:
            latest = {
                "nr": clean(values[0] if len(values) > 0 else None),
                "data": clean(data),
                "lanciatore": clean(lanciatore),
                "cliente": clean(cliente),
                "stato": clean(values[5] if len(values) > 5 else None),
                "orbita": clean(values[7] if len(values) > 7 else None),
                "landing": clean(values[8] if len(values) > 8 else None),
                "booster": clean(values[9] if len(values) > 9 else None),
                "voli": clean(values[10] if len(values) > 10 else None),
                "pad": clean(values[11] if len(values) > 11 else None),
            }
    return {"metrics": metrics, "annual": annual, "launchers": launchers, "pads": pads, "landing": landing, "latest": latest}


def starship_data():
    path = ROOT / "01_workbook" / "sviluppo_starship.xlsx"
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    dashboard = workbook["Dashboard"]
    metrics = {
        "eventi": clean(dashboard["A5"].value),
        "voli": clean(dashboard["C5"].value),
        "catch": clean(dashboard["E5"].value),
        "reflight": clean(dashboard["G5"].value),
        "stato": clean(dashboard["I5"].value),
        "statoDettaglio": clean(dashboard["I6"].value),
    }
    flights = [
        {
            "flight": row.get("Flight"),
            "data": row.get("Data"),
            "milestone": row.get("Milestone"),
            "esito": row.get("Esito integrato"),
        }
        for row in rows_from_sheet(path, "Voli integrati", 4)
        if row.get("Flight")
    ]
    themes = [
        {
            "tema": row.get("Tema"),
            "sintesi": row.get("Sintesi integrata"),
            "impatto": row.get("Impatto"),
        }
        for row in rows_from_sheet(path, "Temi e sintesi", 4)
        if row.get("Tema")
    ]
    return {"metrics": metrics, "flights": flights, "themes": themes}


def literal_assignment(path, name):
    text = path.read_text(encoding="utf-8")
    marker = f"{name} = "
    start = text.index(marker) + len(marker)
    while start < len(text) and text[start].isspace():
        start += 1
    if text[start] not in "[{(":
        raise ValueError(f"{name} literal not found")
    opener = text[start]
    closer = {("["): "]", ("{"): "}", ("("): ")"}[opener]
    depth = 0
    in_string = None
    escaped = False
    for index in range(start, len(text)):
        char = text[index]
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == in_string:
                in_string = None
            continue
        if char in ("'", '"'):
            in_string = char
        elif char == opener:
            depth += 1
        elif char == closer:
            depth -= 1
            if depth == 0:
                return ast.literal_eval(text[start : index + 1])
    raise ValueError(f"{name} literal not closed")


def pad_launch_data():
    coords = {
        "SLC-40": (28.5619, -80.5772, "Florida"),
        "LC-39A": (28.6084, -80.6043, "Florida"),
        "SLC-4E": (34.6320, -120.6106, "California"),
        "SLC-6": (34.5811, -120.6266, "California"),
        "BOCA CHICA PAD 1": (25.9971, -97.1567, "Texas"),
        "BOCA CHICA PAD 2": (25.9980, -97.1544, "Texas"),
    }
    images = {
        "SLC-40": "pad_slc-40.png",
        "LC-39A": "pad_lc-39a.png",
        "SLC-4E": "pad_slc-4e.png",
        "SLC-6": "pad_slc-6.png",
        "BOCA CHICA PAD 1": "pad_boca_chica_pad_1.png",
        "BOCA CHICA PAD 2": "pad_boca_chica_pad_2.png",
    }
    pads = [
        {
            "title": "SLC-37B",
            "subtitle": "Da pad storico Saturn e Delta IV a futuro sito Starship di SpaceX",
            "location": "Cape Canaveral Space Force Station, Florida",
            "full": "Space Launch Complex 37B",
            "lat": 28.5319,
            "lng": -80.5648,
            "region": "Florida",
            "image": "pad 37b.png",
            "timeline": [
                {"date": "1964-1968", "text": "SLC-37 viene usato per missioni Saturn I e Saturn IB durante l'era Apollo."},
                {"date": "2002-2024", "text": "Il complesso viene riattivato come SLC-37B per il programma Delta IV e Delta IV Heavy di ULA."},
                {"date": "9 aprile 2024", "text": "Ultimo lancio Delta IV Heavy dal pad con NROL-70: si chiude l'era Delta IV a SLC-37B."},
                {"date": "21 febbraio 2024", "text": "Il Department of the Air Force avvia l'EIS per valutare SLC-37 come futuro sito Starship/Super Heavy."},
                {"date": "Marzo 2025", "text": "SpaceX ottiene un limited Right of Entry per sopralluoghi, verifiche tecniche e attivita preliminari."},
                {"date": "6 giugno 2025", "text": "Viene pubblicato il Draft EIS: la proposta prevede la ricostruzione del sito per supportare fino a 76 lanci Starship/Super Heavy all'anno."},
                {"date": "12 giugno 2025", "text": "Demolizione della storica infrastruttura Delta IV, inclusa la Mobile Service Tower. Inizia la trasformazione fisica del complesso."},
                {"date": "20 novembre 2025", "text": "Il Department of the Air Force firma il Record of Decision: SpaceX e autorizzata a riqualificare SLC-37 per Starship/Super Heavy tramite accordi immobiliari con la U.S. Space Force."},
                {"date": "Dicembre 2025", "text": "SpaceX comunica che l'approvazione e stata ricevuta e che la costruzione e iniziata."},
            ],
            "current": [
                "SLC-37B e nella fase di assegnazione e riqualificazione a favore di SpaceX.",
                "Non si tratta di un normale affitto pubblico gia visibile come contratto privato: la forma ufficiale indicata nei documenti e un real property agreement con la U.S. Space Force.",
                "La riconversione del sito e in corso.",
                "Prima dei lanci operativi Starship da questo pad serviranno ulteriori passaggi regolatori FAA, in particolare sulle analisi dello spazio aereo.",
                "Capacita prevista: fino a 76 lanci e 152 atterraggi annui.",
            ],
            "source_note": "Sintesi cronologica basata su documentazione pubblica DAF/USSF, processo EIS e comunicazioni SpaceX.",
        }
    ]

    generated = literal_assignment(ROOT / "03_script" / "genera_infografiche_pad.py", "PADS")
    for item in generated:
        lat, lng, region = coords[item["title"]]
        pads.append(
            {
                "title": item["title"],
                "subtitle": item["subtitle"],
                "location": item["location"],
                "full": item["full"],
                "lat": lat,
                "lng": lng,
                "region": region,
                "image": images[item["title"]],
                "timeline": item["timeline"],
                "current": item["current"],
                "source_note": item["source_note"],
            }
        )
    for index, item in enumerate(pads, start=1):
        item["id"] = f"pad-{index}"
        item["imageHref"] = f"../pad_di_lancio/{quote(item['image'])}"
    return pads


def spacex_locations_data():
    return [
        {
            "id": "loc-hawthorne",
            "name": "SpaceX Hawthorne",
            "short": "Hawthorne, California",
            "region": "California",
            "status": "Operativa",
            "role": "Sede storica e produzione",
            "lat": 33.9199,
            "lng": -118.3275,
            "address": "1 Rocket Road, Hawthorne, CA 90250",
            "coords": "33.9199 N, 118.3275 W",
            "summary": "Sede storica e principale stabilimento ingegneristico-produttivo di SpaceX.",
            "details": [
                "Qui sono stati progettati e costruiti Falcon 9, Falcon Heavy e Dragon.",
                "Resta un nodo tecnico e produttivo centrale anche dopo gli annunci di trasferimento operativo in Texas.",
                "Il marker rappresenta il complesso di Rocket Road.",
            ],
            "source_note": "Punto e descrizione ripresi dal file mappa locale originario.",
        },
        {
            "id": "loc-redmond",
            "name": "SpaceX Starlink Redmond",
            "short": "Redmond, Washington",
            "region": "Washington",
            "status": "Operativa",
            "role": "Produzione satelliti Starlink",
            "lat": 47.6939,
            "lng": -122.0336,
            "address": "Redmond, Washington, area Seattle",
            "coords": "47.6939 N, 122.0336 W",
            "summary": "Stabilimento dedicato alla produzione in serie dei satelliti Starlink.",
            "details": [
                "La struttura sostiene il rapido dispiegamento della costellazione Starlink.",
                "Il punto e rappresentativo dell'area industriale Redmond usata per Starlink.",
                "La mappa locale indicava una media 2026 di circa 70 satelliti prodotti a settimana.",
            ],
            "source_note": "Punto rappresentativo e note dal file mappa locale originario.",
        },
        {
            "id": "loc-vandenberg",
            "name": "SpaceX Vandenberg SLC-4E",
            "short": "Vandenberg SFB, California",
            "region": "California",
            "status": "Operativa",
            "role": "Lanci costa ovest",
            "lat": 34.6320,
            "lng": -120.6106,
            "address": "Space Launch Complex 4 East, Vandenberg Space Force Base, CA",
            "coords": "34.6320 N, 120.6106 W",
            "summary": "Pad Falcon 9 per orbite polari, SSO e missioni dalla costa ovest.",
            "details": [
                "Usato per missioni Starlink ad alta inclinazione e payload governativi.",
                "Lavora con LZ-4 per i ritorni a terra quando il profilo missione lo consente.",
                "E il riferimento operativo SpaceX principale sulla West Coast.",
            ],
            "source_note": "Coordinate e descrizione dalla mappa locale, coerenti con la pagina pad.",
        },
        {
            "id": "loc-mcgregor",
            "name": "SpaceX McGregor Test Facility",
            "short": "McGregor, Texas",
            "region": "Texas",
            "status": "Operativa",
            "role": "Test motori e stadi",
            "lat": 31.3916,
            "lng": -97.4574,
            "address": "McGregor Rocket Development and Test Facility, McGregor, TX",
            "coords": "31.3916 N, 97.4574 W",
            "summary": "Complesso di prova per motori, propulsori e stadi prima del volo.",
            "details": [
                "Qui vengono collaudati motori Merlin e Raptor.",
                "La struttura ospita prove statiche e test di qualifica di componenti critici.",
                "Il punto indica l'area rappresentativa del grande sito di test.",
            ],
            "source_note": "Punto e sintesi derivati dal file mappa locale originario.",
        },
        {
            "id": "loc-starbase",
            "name": "SpaceX Starbase",
            "short": "Starbase, Texas",
            "region": "Texas",
            "status": "Operativa / in espansione",
            "role": "Sviluppo, produzione e lancio Starship",
            "lat": 25.9970,
            "lng": -97.1570,
            "address": "Boca Chica Blvd, Starbase, Brownsville, TX 78521",
            "coords": "25.9970 N, 97.1570 W",
            "summary": "Sito integrato di sviluppo, produzione e lancio orbitale Starship.",
            "details": [
                "Comprende fabbriche, pad orbitali, aree di test e infrastrutture di controllo.",
                "E il centro piu visibile della transizione SpaceX verso Starship.",
                "Il marker e un punto rappresentativo dell'area Starbase/Boca Chica.",
            ],
            "source_note": "Punto rappresentativo dal file mappa locale originario.",
        },
        {
            "id": "loc-lc39a",
            "name": "SpaceX LC-39A",
            "short": "Kennedy Space Center, Florida",
            "region": "Florida",
            "status": "Operativa",
            "role": "Crew Dragon, Falcon Heavy, missioni speciali",
            "lat": 28.6084,
            "lng": -80.6043,
            "address": "Launch Complex 39A, Kennedy Space Center, Merritt Island, FL",
            "coords": "28.6084 N, 80.6043 W",
            "summary": "Pad storico Apollo/Shuttle gestito da SpaceX per missioni ad alta complessita.",
            "details": [
                "Usato per Crew Dragon, Falcon Heavy e missioni commerciali o scientifiche speciali.",
                "Resta uno dei pad piu iconici e strategici della Space Coast.",
                "La pagina pad contiene la cronologia infrastrutturale completa.",
            ],
            "source_note": "Coordinate e testo sintetico dal file mappa locale originario.",
        },
        {
            "id": "loc-slc40",
            "name": "SpaceX SLC-40",
            "short": "Cape Canaveral SFS, Florida",
            "region": "Florida",
            "status": "Operativa",
            "role": "Workhorse Falcon 9",
            "lat": 28.5619,
            "lng": -80.5772,
            "address": "Space Launch Complex 40, Cape Canaveral Space Force Station, FL",
            "coords": "28.5619 N, 80.5772 W",
            "summary": "Pad Falcon 9 ad altissima cadenza sulla costa est.",
            "details": [
                "Supporta gran parte dei lanci Starlink, CRS e payload commerciali.",
                "La torre Dragon ha aumentato la ridondanza rispetto a LC-39A.",
                "E il sito operativo piu regolare nella macchina Falcon 9.",
            ],
            "source_note": "Coordinate e sintesi dal file mappa locale originario.",
        },
        {
            "id": "loc-louisiana",
            "name": "SpaceX Louisiana",
            "short": "Pecan Island / Freshwater City, Louisiana",
            "region": "Louisiana",
            "status": "Proposta / ipotesi",
            "role": "Possibile futuro mega-sito Starship",
            "lat": 29.580,
            "lng": -92.450,
            "address": "Zona costiera a sud della LA-82, Vermilion Parish",
            "coords": "29.580 N, 92.450 W, punto rappresentativo",
            "summary": "Ipotesi di acquisizione, non struttura SpaceX attiva.",
            "details": [
                "Area indicativa di circa 136.000 acri di marshland costiera.",
                "Le note locali parlano di negoziati o interesse per espansione Starship, senza sito operativo avviato.",
                "Nessuna costruzione, licenza FAA o pad operativo e indicato nel file sorgente.",
                "La mappa mostra anche un poligono approssimativo per ricordare che il punto non e una localita puntuale.",
            ],
            "source_note": "Ipotesi riportata nel file locale con fonti giornalistiche e dichiarazioni locali citate.",
            "proposed": True,
            "polygon": [
                [29.632, -92.525],
                [29.632, -92.365],
                [29.385, -92.365],
                [29.385, -92.525],
            ],
        },
    ]


def json_for_html(data):
    return json.dumps(data, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")


def page_path(slug):
    return f"{slug}.html"


def section_href(slug, from_section):
    return page_path(slug) if from_section else f"sezioni/{page_path(slug)}"


def nav_html(current, from_section):
    home_href = "../index.html" if from_section else "index.html"
    links = [("Home", home_href)]
    links.extend((item["title"], section_href(item["slug"], from_section)) for item in MAIN_SECTIONS)
    return "\n".join(
        f'<a class="{ "active" if key == current else "" }" href="{href}">{escape(label)}</a>'
        for label, href in links
        for key in [label.lower().replace(" ", "-") if label != "Home" else "home"]
    )


def head_html(title, css_href, extra_head=""):
    return f"""<!doctype html>
<html lang="it">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{escape(title)} | Rivoluzione Spaziale</title>
<link rel="stylesheet" href="{css_href}">
{extra_head}
</head>
"""


def shell(title, current, from_section, body, extra_script="", extra_head=""):
    css_path = "../css/style.css" if from_section else "css/style.css"
    css_href = f"{css_path}?v={CSS_VERSION}"
    return f"""{head_html(title, css_href, extra_head)}
<body>
<header class="topbar">
  <a class="brand" href="{'../index.html' if from_section else 'index.html'}">Rivoluzione <span>Spaziale</span></a>
  <nav class="nav" aria-label="Navigazione principale">
{nav_html(current, from_section)}
  </nav>
</header>
<main>
{body}
</main>
<footer class="footer">
  Rivoluzione Spaziale. Sito statico multipagina pubblicato su GitHub Pages.
</footer>
{extra_script}
</body>
</html>
"""


def metric(value, label):
    return f'<div class="metric"><b>{escape(str(value))}</b><span>{escape(label)}</span></div>'


def percent(value):
    if value is None:
        return "n.d."
    return f"{round(value * 1000) / 10}%"


def display_date(value):
    if not value:
        return "n.d."
    try:
        parsed = datetime.fromisoformat(str(value))
        return parsed.strftime("%d/%m/%Y")
    except ValueError:
        return str(value)


def render_css():
    return f""":root {{
  --bg:#050607;
  --ink:#f5f7fa;
  --muted:#aeb6bd;
  --line:rgba(255,255,255,.16);
  --panel:rgba(255,255,255,.07);
  --panel2:rgba(255,255,255,.105);
  --accent:#69c8ff;
  --gold:#f2b84b;
  --green:#78d58c;
}}
*,*::before,*::after{{box-sizing:border-box}}
html{{scroll-behavior:smooth;overflow-x:clip}}
body{{margin:0;background:var(--bg);color:var(--ink);font-family:Inter,Segoe UI,Roboto,Arial,sans-serif;letter-spacing:0;overflow-x:hidden}}
a{{color:inherit;text-decoration:none}}
.topbar{{position:sticky;top:0;z-index:30;display:flex;align-items:center;justify-content:space-between;gap:18px;padding:16px clamp(18px,4vw,56px);background:rgba(5,6,7,.92);border-bottom:1px solid var(--line);backdrop-filter:blur(12px)}}
.brand{{font-weight:950;letter-spacing:.18em;text-transform:uppercase;font-size:14px;white-space:nowrap}}
.brand span{{color:var(--accent)}}
.nav{{display:flex;gap:12px;flex-wrap:wrap;justify-content:flex-end;font-size:11px;text-transform:uppercase;letter-spacing:.08em;font-weight:850}}
.nav a{{color:#fff;opacity:.72}}
.nav a:hover,.nav a.active{{opacity:1;color:#fff}}
.hero{{min-height:76vh;display:grid;align-items:end;padding:110px clamp(18px,4vw,56px) 52px;background:linear-gradient(90deg,rgba(0,0,0,.92),rgba(0,0,0,.54) 45%,rgba(0,0,0,.12)),linear-gradient(180deg,rgba(0,0,0,.08),rgba(5,6,7,.96) 92%),url('{HERO_IMAGE}') center/cover no-repeat}}
.page-hero{{padding:76px clamp(18px,4vw,56px) 44px;border-bottom:1px solid var(--line);background:linear-gradient(180deg,rgba(255,255,255,.07),rgba(255,255,255,.02))}}
.hero-inner,.inner{{max-width:1240px;margin:0 auto}}
.eyebrow{{margin:0 0 14px;color:var(--gold);font-size:13px;font-weight:900;text-transform:uppercase;letter-spacing:.15em}}
h1{{margin:0;max-width:1000px;font-size:clamp(42px,8vw,106px);line-height:.9;text-transform:uppercase;font-weight:950}}
h2{{margin:0;font-size:clamp(30px,4vw,56px);line-height:.98;text-transform:uppercase}}
h3{{margin:0 0 10px;font-size:22px;line-height:1.14}}
p{{color:#d9dee3;line-height:1.64}}
.subtitle{{max-width:820px;margin:24px 0 0;color:#e6ebef;font-size:clamp(18px,2vw,24px);line-height:1.45;font-weight:560}}
.actions{{display:flex;gap:12px;flex-wrap:wrap;margin-top:30px}}
.button{{border:1px solid var(--line);padding:13px 18px;border-radius:4px;text-transform:uppercase;font-size:12px;letter-spacing:.09em;font-weight:900;background:#fff;color:#000}}
.button.secondary{{background:rgba(0,0,0,.22);color:#fff}}
.spacex-actions{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));max-width:430px}}
.spacex-actions .button{{display:flex;align-items:center;justify-content:center;text-align:center;min-height:44px;padding:12px 14px}}
.spacex-actions .button.history-link{{grid-column:1/-1}}
section{{padding:72px clamp(18px,4vw,56px);border-top:1px solid var(--line)}}
.section-head{{display:flex;justify-content:space-between;gap:28px;align-items:end;margin-bottom:28px}}
.section-head p{{max-width:660px;color:var(--muted);line-height:1.58;margin:0}}
.grid{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:14px}}
.card,.panel{{border:1px solid var(--line);background:linear-gradient(180deg,var(--panel2),var(--panel));border-radius:8px;padding:20px}}
.card{{min-height:182px;display:flex;flex-direction:column;justify-content:space-between;transition:transform .18s ease,border-color .18s ease}}
.card:hover{{transform:translateY(-3px);border-color:rgba(105,200,255,.55)}}
.card small,.badge{{color:var(--gold);text-transform:uppercase;letter-spacing:.1em;font-size:11px;font-weight:900}}
.card p{{margin:0;color:#cbd2d8;font-size:14px}}
.split{{display:grid;grid-template-columns:1.05fr .95fr;gap:20px;align-items:stretch}}
.metrics{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}}
.split .panel .metrics{{grid-template-columns:repeat(2,minmax(0,1fr));margin-top:12px}}
.metric{{min-width:0;border:1px solid var(--line);background:rgba(255,255,255,.065);border-radius:8px;padding:18px;min-height:112px}}
.metric b{{display:block;font-size:clamp(25px,2.6vw,34px);line-height:1.05;margin-bottom:8px;overflow-wrap:break-word}}
.metric span{{display:block;color:var(--muted);font-size:13px;line-height:1.35}}
.story-list{{display:grid;gap:16px}}
.story-card{{display:grid;grid-template-columns:minmax(220px,320px) minmax(0,1fr);gap:20px;align-items:stretch;border:1px solid var(--line);background:linear-gradient(180deg,var(--panel2),var(--panel));border-radius:8px;padding:14px;transition:transform .18s ease,border-color .18s ease}}
.story-card:hover{{transform:translateY(-3px);border-color:rgba(105,200,255,.55)}}
.story-card img{{display:block;width:100%;height:100%;min-height:230px;object-fit:cover;border-radius:6px;border:1px solid rgba(255,255,255,.16);background:#090d11}}
.story-card small{{display:block;color:var(--gold);text-transform:uppercase;letter-spacing:.1em;font-size:11px;font-weight:900;margin-bottom:10px}}
.story-card p{{max-width:760px;margin:10px 0 0;color:#cbd2d8}}
.story-card .button{{display:inline-flex;margin-top:18px}}
.agenda-strip{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:12px;margin-top:22px}}
.agenda-note{{border:1px solid var(--line);background:rgba(255,255,255,.055);border-radius:8px;padding:16px;min-height:96px}}
.agenda-note b{{display:block;font-size:24px;margin-bottom:6px}}
.agenda-note span{{display:block;color:var(--muted);font-size:13px;line-height:1.4}}
.next-launch{{display:grid;grid-template-columns:minmax(0,.9fr) minmax(0,1.1fr);gap:18px;border:1px solid rgba(105,200,255,.45);background:linear-gradient(135deg,rgba(105,200,255,.16),rgba(255,255,255,.055));border-radius:8px;padding:20px;margin-top:18px}}
.next-launch h3{{font-size:30px}}
.next-launch p{{margin:8px 0 0;color:#d9e2e8}}
.launch-grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}}
.launch-grid.compact{{grid-template-columns:repeat(3,minmax(0,1fr))}}
.launch{{border:1px solid var(--line);background:linear-gradient(180deg,rgba(255,255,255,.1),rgba(255,255,255,.05));border-radius:8px;padding:16px;display:flex;flex-direction:column;gap:12px;min-height:360px}}
.launch-top{{display:flex;justify-content:space-between;gap:10px;align-items:flex-start}}
.pill{{display:inline-flex;border:1px solid rgba(105,200,255,.5);color:#dff4ff;border-radius:999px;padding:5px 8px;font-size:11px;text-transform:uppercase;letter-spacing:.08em;font-weight:900}}
.pill.net{{border-color:rgba(242,184,75,.6);color:#ffe1a3}}
.date{{color:#d6dde3;text-align:right;font-size:13px;font-weight:800}}
.launch h3{{font-size:21px}}
.launch p{{margin:0;font-size:14px;color:#cbd2d8}}
.launch-summary{{display:grid;gap:8px}}
.launch-summary p{{line-height:1.45}}
.launch-meta{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:8px}}
.launch-meta div{{border:1px solid rgba(255,255,255,.12);border-radius:6px;padding:9px;background:rgba(0,0,0,.2);min-width:0}}
.launch-meta b{{display:block;color:var(--gold);font-size:10px;text-transform:uppercase;letter-spacing:.08em;margin-bottom:5px}}
.launch-meta span{{display:block;color:#dce4ea;font-size:12px;line-height:1.35;overflow-wrap:break-word}}
.history-meta{{grid-template-columns:repeat(4,minmax(0,1fr));margin-top:12px}}
.source-links{{display:flex;flex-wrap:wrap;gap:7px;margin-top:2px}}
.source-links a{{border:1px solid rgba(255,255,255,.16);border-radius:999px;padding:5px 8px;color:#dcebf4;font-size:11px;text-transform:uppercase;letter-spacing:.06em;font-weight:850}}
.countdown{{margin-top:auto;display:grid;grid-template-columns:repeat(4,1fr);gap:6px}}
.countdown.pending{{grid-template-columns:1fr}}
.timebox{{background:rgba(0,0,0,.28);border:1px solid rgba(255,255,255,.1);border-radius:6px;padding:8px 4px;text-align:center}}
.timebox strong{{display:block;font-size:20px}}
.timebox span{{display:block;color:var(--muted);font-size:10px;text-transform:uppercase;letter-spacing:.08em}}
.dash-grid{{display:grid;grid-template-columns:1.2fr .8fr;gap:18px}}
.cols{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:18px}}
.bars{{display:grid;gap:8px}}
.bar-row{{display:grid;grid-template-columns:72px 1fr 56px;gap:10px;align-items:center;font-size:13px;color:#dfe5ea}}
.bar{{height:9px;background:rgba(255,255,255,.1);border-radius:999px;overflow:hidden}}
.bar i{{display:block;height:100%;background:linear-gradient(90deg,var(--accent),var(--green));border-radius:999px}}
table{{width:100%;border-collapse:collapse;font-size:14px}}
th,td{{border-top:1px solid var(--line);padding:12px 8px;text-align:left;vertical-align:top;line-height:1.38}}
th{{color:#fff;text-transform:uppercase;letter-spacing:.08em;font-size:11px}}
td{{color:#d8dee3}}
.theme-list{{display:grid;gap:12px}}
.theme{{border-left:3px solid var(--accent);padding:0 0 0 14px}}
.theme p{{margin:4px 0 0;font-size:14px;color:#cfd6dc}}
.pad-map-wrap{{display:grid;grid-template-columns:320px 1fr;min-height:620px;border:1px solid var(--line);border-radius:8px;overflow:hidden;background:#080b0f}}
.pad-map-side{{background:rgba(255,255,255,.055);border-right:1px solid var(--line);padding:18px;display:flex;flex-direction:column;gap:12px}}
.pad-map-side h3{{margin-bottom:2px}}
.pad-map-side p{{font-size:14px;color:var(--muted);margin:0}}
.pad-filter-row{{display:flex;flex-wrap:wrap;gap:8px;margin-top:6px}}
.pad-filter{{border:1px solid var(--line);background:rgba(0,0,0,.2);color:#fff;border-radius:999px;padding:8px 10px;font-size:11px;text-transform:uppercase;letter-spacing:.08em;font-weight:900;cursor:pointer}}
.pad-filter.active{{background:#fff;color:#000;border-color:#fff}}
.pad-side-list{{display:grid;gap:7px;margin-top:8px;overflow:auto;max-height:410px;padding-right:3px}}
.pad-side-item{{border:1px solid transparent;border-radius:8px;background:rgba(255,255,255,.045);padding:10px;text-align:left;color:#fff;cursor:pointer}}
.pad-side-item:hover,.pad-side-item.active{{border-color:rgba(105,200,255,.58);background:rgba(105,200,255,.1)}}
.pad-side-item strong{{display:block;font-size:14px;line-height:1.2}}
.pad-side-item span{{display:block;color:var(--muted);font-size:12px;line-height:1.35;margin-top:3px}}
#pad-map,#location-map{{min-height:620px;background:#07101a}}
.leaflet-container{{font-family:Inter,Segoe UI,Roboto,Arial,sans-serif;background:#07101a}}
.leaflet-popup-content-wrapper{{background:#101418;color:var(--ink);border-radius:8px;border:1px solid rgba(255,255,255,.16);box-shadow:0 18px 50px rgba(0,0,0,.42);overflow:hidden}}
.leaflet-popup-content{{margin:0!important;line-height:1.45}}
.leaflet-popup-tip{{background:#101418}}
.leaflet-control-zoom a,.leaflet-control-layers{{background:#111820!important;color:#e8edf2!important;border-color:rgba(255,255,255,.18)!important}}
.pad-marker{{background:transparent;border:0}}
.pad-marker-dot{{width:34px;height:34px;border-radius:50%;display:grid;place-items:center;background:var(--accent);color:#051018;border:3px solid #fff;font-size:10px;font-weight:950;box-shadow:0 6px 18px rgba(0,0,0,.35)}}
.pad-popup{{width:min(360px,78vw);background:#101418}}
.pad-popup img{{display:block;width:100%;height:158px;object-fit:cover;object-position:top;border-bottom:1px solid var(--line)}}
.pad-popup-body{{padding:14px}}
.pad-popup .badge{{display:inline-block;margin-bottom:8px;color:var(--gold)}}
.pad-popup h3{{font-size:22px;margin:0 0 4px}}
.pad-popup p{{font-size:13px;color:#d7dee4;line-height:1.45;margin:8px 0}}
.pad-popup ul{{margin:8px 0 0;padding-left:18px;color:#d7dee4;font-size:13px;line-height:1.38}}
.pad-popup li{{margin-bottom:4px}}
.location-popup{{width:min(420px,calc(100vw - 52px));max-height:min(640px,78vh);overflow:auto;background:#101418}}
.location-popup-head{{padding:14px 14px 12px;border-bottom:1px solid rgba(255,255,255,.12);background:linear-gradient(135deg,rgba(105,200,255,.18),rgba(255,255,255,.045))}}
.location-popup.proposed .location-popup-head{{background:linear-gradient(135deg,rgba(242,184,75,.24),rgba(255,255,255,.045))}}
.location-popup h3{{font-size:21px;margin:7px 0 4px;line-height:1.12}}
.location-popup p{{font-size:13px;color:#d8e1e6;line-height:1.45;margin:0}}
.location-popup-body{{padding:14px;display:grid;gap:10px}}
.location-popup-box{{border:1px solid rgba(255,255,255,.12);border-radius:7px;background:rgba(0,0,0,.26);padding:11px}}
.location-popup-box b{{display:block;color:var(--gold);text-transform:uppercase;letter-spacing:.08em;font-size:10px;margin-bottom:5px}}
.location-popup-box ul{{margin:0;padding-left:17px;color:#d8e1e6;font-size:13px;line-height:1.42}}
.location-popup-box li{{margin:0 0 5px}}
.location-list-item{{width:100%;text-align:left;border:1px solid rgba(255,255,255,.12);background:rgba(255,255,255,.045);border-radius:7px;padding:11px;color:#fff;cursor:pointer}}
.location-list-item:hover,.location-list-item.active{{border-color:rgba(105,200,255,.55);background:rgba(105,200,255,.12)}}
.location-list-item.proposed:hover,.location-list-item.proposed.active{{border-color:rgba(242,184,75,.65);background:rgba(242,184,75,.13)}}
.location-list-item strong{{display:block;font-size:14px;line-height:1.25}}
.location-list-item span{{display:block;color:var(--muted);font-size:12px;line-height:1.32;margin-top:4px}}
.location-dot{{display:inline-flex;align-items:center;justify-content:center;width:34px;height:34px;border-radius:999px;border:2px solid #fff;background:#d34949;color:#fff;font-size:11px;font-weight:950;box-shadow:0 8px 18px rgba(0,0,0,.35)}}
.location-dot.proposed{{border-radius:8px;border-style:dashed;background:#d99b28;color:#130d04}}
.pad-list{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:16px}}
.pad-card{{display:grid;grid-template-columns:220px 1fr;gap:18px;border:1px solid var(--line);border-radius:8px;background:linear-gradient(180deg,var(--panel2),var(--panel));padding:14px;min-width:0}}
.pad-card figure{{margin:0;min-width:0}}
.pad-card img{{width:100%;aspect-ratio:1055/1491;object-fit:cover;object-position:top;border-radius:6px;border:1px solid rgba(255,255,255,.16);background:#0a0f14}}
.pad-card h3{{font-size:28px}}
.pad-card p{{font-size:14px;margin:8px 0;color:#cfd7de}}
.pad-card dl{{display:grid;grid-template-columns:auto 1fr;gap:6px 10px;margin:10px 0 0;color:#dce4ea;font-size:13px}}
.pad-card dt{{color:var(--muted);text-transform:uppercase;letter-spacing:.08em;font-size:10px;font-weight:900}}
.pad-card dd{{margin:0;min-width:0}}
.pad-mini-timeline{{margin-top:12px;display:grid;gap:8px}}
.pad-mini-event{{display:grid;grid-template-columns:104px 1fr;gap:10px;border-top:1px solid rgba(255,255,255,.12);padding-top:8px;font-size:13px;color:#d9e0e6;line-height:1.38}}
.pad-mini-event strong{{color:var(--accent);font-size:12px;line-height:1.25}}
.pad-current-list{{margin:12px 0 0;padding-left:18px;color:#d7dee4;font-size:13px;line-height:1.45}}
.pad-current-list li{{margin-bottom:6px}}
.source-panel{{border:1px solid var(--line);border-radius:8px;background:rgba(255,255,255,.055);padding:18px}}
.source-panel p{{margin:0 0 10px;color:var(--muted);font-size:14px}}
.source-panel a{{color:#dff4ff;text-decoration:underline;text-underline-offset:3px}}
.construction{{border-color:rgba(242,184,75,.45);background:linear-gradient(180deg,rgba(242,184,75,.12),rgba(255,255,255,.055))}}
.muted{{color:var(--muted)}}
.footer{{padding:34px clamp(18px,4vw,56px);border-top:1px solid var(--line);color:var(--muted);font-size:13px;line-height:1.5}}
@media(max-width:1120px){{.grid,.launch-grid,.launch-grid.compact,.agenda-strip,.pad-list{{grid-template-columns:repeat(2,minmax(0,1fr))}}.split,.dash-grid,.cols,.next-launch,.pad-map-wrap{{grid-template-columns:1fr}}.pad-map-side{{border-right:0;border-bottom:1px solid var(--line)}}.pad-side-list{{max-height:none;grid-template-columns:repeat(2,minmax(0,1fr))}}}}
@media(max-width:900px){{.pad-list{{grid-template-columns:1fr}}.pad-card{{grid-template-columns:160px 1fr}}.story-card{{grid-template-columns:1fr}}.story-card img{{height:auto;max-height:420px}}}}
@media(max-width:760px){{.topbar{{align-items:flex-start;flex-direction:column}}.nav{{justify-content:flex-start}}.grid,.launch-grid,.launch-grid.compact,.agenda-strip,.launch-meta,.metrics,.split .panel .metrics,.pad-side-list,.spacex-actions{{grid-template-columns:1fr}}section{{padding:56px 18px}}.hero{{padding:92px 18px 42px}}h1{{font-size:43px}}.section-head{{display:block}}.bar-row{{grid-template-columns:58px 1fr 44px}}#pad-map,#location-map{{min-height:460px}}.pad-map-wrap{{min-height:460px}}.pad-card{{grid-template-columns:1fr}}.pad-card img{{max-height:420px}}.pad-mini-event{{grid-template-columns:1fr;gap:2px}}}}
"""


def render_home():
    cards = "\n".join(
        f"""<a class="card {'construction' if not item.get('active') else ''}" href="sezioni/{item['slug']}.html">
  <small>{escape(item['status'])}</small>
  <div><h3>{escape(item['title'])}</h3><p>{escape(item['copy'])}</p></div>
</a>"""
        for item in MAIN_SECTIONS
    )
    body = f"""
<section class="hero">
  <div class="hero-inner">
    <p class="eyebrow">Osservatorio</p>
    <h1>Rivoluzione Spaziale</h1>
    <p class="subtitle">I nuovi signori dello spazio. Una mappa editoriale per seguire chi sta ridisegnando accesso all'orbita, cadenza industriale e competizione tra compagnie di lancio.</p>
    <div class="actions">
      <a class="button" href="sezioni/spacex.html">Apri SpaceX</a>
      <a class="button secondary" href="#sezioni-principali">Sezioni principali</a>
    </div>
  </div>
</section>
<section id="sezioni-principali">
  <div class="inner">
    <div class="section-head">
      <h2>Sezioni principali</h2>
      <p>La pagina principale mostra solo il primo livello del sito: compagnie e aree industriali. Le pagine operative, come lanci, Starship e pad di lancio, vivono dentro la sezione della compagnia a cui appartengono.</p>
    </div>
    <div class="grid">{cards}</div>
  </div>
</section>
"""
    return shell("Home", "home", False, body)


def render_launch_cards(launches, compact=False):
    cards = []
    for item in launches:
        categories = item.get("cat") or []
        is_net = "net" in categories
        meta_rows = "".join(
            f"<div><b>{escape(label)}</b><span>{escape(str(value))}</span></div>"
            for label, value in [
                ("Finestra", item.get("window") or item.get("dateLabel") or "Da confermare"),
                ("Orbita", item.get("orbit") or "Da confermare"),
                ("Recupero", item.get("landing") or "Da confermare"),
            ]
        )
        sources = "".join(
            f'<a href="{escape(str(src.get("url") or "#"))}" target="_blank" rel="noopener">{escape(str(src.get("label") or "Fonte"))}</a>'
            for src in item.get("sources", [])
        )
        countdown = (
            f"""<div class="countdown pending">
    <div class="timebox"><strong>NET</strong><span>{escape(str(item.get('dateLabel') or 'finestra indicativa'))}</span></div>
  </div>"""
            if is_net
            else """<div class="countdown">
    <div class="timebox"><strong class="d">0</strong><span>giorni</span></div>
    <div class="timebox"><strong class="h">00</strong><span>ore</span></div>
    <div class="timebox"><strong class="m">00</strong><span>min</span></div>
    <div class="timebox"><strong class="s">00</strong><span>sec</span></div>
  </div>"""
        )
        cards.append(
            f"""<article class="launch" data-iso="{escape(str(item.get('iso') or ''))}">
  <div class="launch-top">
    <span class="pill {'net' if is_net else ''}">{'NET' if is_net else 'T-0'}</span>
    <span class="date">{escape(str(item.get('dateLabel') or 'Data da confermare'))}</span>
  </div>
  <h3>{escape(str(item.get('name') or 'Missione'))}</h3>
  <p><strong>{escape(str(item.get('rocket') or ''))}</strong></p>
  <p>{escape(str(item.get('site') or ''))}</p>
  <p>{escape(str(item.get('payload') or ''))}</p>
  <div class="launch-meta">{meta_rows}</div>
  <div class="launch-summary">
    <p>{escape(str(item.get('summary') or ''))}</p>
    <p class="muted">{escape(str(item.get('news') or item.get('status') or ''))}</p>
  </div>
  <div class="source-links">{sources}</div>
  {countdown}
</article>"""
        )
    class_name = "launch-grid compact" if compact else "launch-grid"
    return f'<div class="{class_name}">' + "\n".join(cards) + "</div>"


def countdown_script():
    return """<script>
function pad(n){return String(n).padStart(2,'0')}
function countdown(iso){
  const diff = new Date(iso).getTime() - Date.now();
  if (!Number.isFinite(diff) || diff <= 0) return ['0','00','00','00'];
  const total = Math.floor(diff / 1000);
  return [Math.floor(total / 86400), pad(Math.floor(total % 86400 / 3600)), pad(Math.floor(total % 3600 / 60)), pad(total % 60)];
}
function tick(){
  document.querySelectorAll('.launch[data-iso]').forEach(card => {
    if (!card.querySelector('.d')) return;
    const parts = countdown(card.dataset.iso);
    card.querySelector('.d').textContent = parts[0];
    card.querySelector('.h').textContent = parts[1];
    card.querySelector('.m').textContent = parts[2];
    card.querySelector('.s').textContent = parts[3];
  });
}
tick();
setInterval(tick, 1000);
</script>"""


def render_falcon_dashboard(falcon):
    f = falcon["metrics"]
    latest = falcon.get("latest") or {}
    latest_meta = "".join(
        f"<div><b>{escape(label)}</b><span>{escape(str(value or 'n.d.'))}</span></div>"
        for label, value in [
            ("Pad", latest.get("pad")),
            ("Orbita", latest.get("orbita")),
            ("Recupero", latest.get("landing")),
            ("Booster", f"{latest.get('booster') or 'n.d.'}, volo {latest.get('voli') or 'n.d.'}"),
        ]
    )
    latest_block = f"""
<article class="next-launch history-latest">
  <div>
    <p class="badge">Ultimo lancio registrato</p>
    <h3>{escape(str(latest.get('cliente') or 'n.d.'))}</h3>
    <p>{escape(display_date(latest.get('data')))} · volo Falcon #{escape(str(latest.get('nr') or 'n.d.'))}</p>
  </div>
  <div>
    <p><strong>{escape(str(latest.get('lanciatore') or 'n.d.'))}</strong></p>
    <p>Esito: {escape(str(latest.get('stato') or 'n.d.'))}</p>
    <div class="launch-meta history-meta">{latest_meta}</div>
  </div>
</article>
"""
    metric_html = "".join(
        [
            metric(f"{f['lanci']:,}".replace(",", "."), "lanci principali"),
            metric(f"{f['successi']:,}".replace(",", "."), "successi"),
            metric(percent(f["tasso"]), "tasso successo"),
            metric(f["boosterMax"], "max voli booster"),
            metric(f["falconHeavy"], "Falcon Heavy"),
            metric(f["padAttivi"], "pad attivi"),
            metric(f["recuperiRiusciti"], "recuperi booster"),
            metric(f["boosterRiutilizzati"], "booster riutilizzati"),
        ]
    )
    max_launches = max(row["lanci"] or 0 for row in falcon["annual"]) or 1
    bars = "\n".join(
        f"""<div class="bar-row"><span>{row['anno']}</span><div class="bar"><i style="width:{max(2, (row['lanci'] or 0) / max_launches * 100)}%"></i></div><strong>{row['lanci']}</strong></div>"""
        for row in falcon["annual"]
    )
    launcher_rows = "\n".join(
        f"<tr><td>{escape(str(row['lanciatore']))}</td><td>{row['lanci']}</td><td>{percent(row['tasso'])}</td></tr>"
        for row in falcon["launchers"]
    )
    pad_rows = "\n".join(
        f"<tr><td>{escape(str(row['pad']))}</td><td>{row['lanci']}</td><td>{percent(row['quota'])}</td></tr>"
        for row in falcon["pads"]
    )
    landing_rows = "\n".join(
        f"<tr><td>{escape(str(row['codice']))}</td><td>{row['record']}</td><td>{row['recuperi']}</td></tr>"
        for row in falcon["landing"]
    )
    return f"""
{latest_block}
<div style="margin-top:18px">
<div class="metrics">{metric_html}</div>
</div>
<div class="dash-grid" style="margin-top:18px">
  <div class="panel"><h3>Lanci per anno</h3><div class="bars">{bars}</div></div>
  <div class="panel"><h3>Famiglie Falcon</h3><table><thead><tr><th>Lanciatore</th><th>Lanci</th><th>Successo</th></tr></thead><tbody>{launcher_rows}</tbody></table></div>
</div>
<div class="cols" style="margin-top:18px">
  <div class="panel"><h3>Pad e aree</h3><table><thead><tr><th>Pad</th><th>Lanci</th><th>Quota</th></tr></thead><tbody>{pad_rows}</tbody></table></div>
  <div class="panel"><h3>Recupero booster</h3><table><thead><tr><th>Landing</th><th>Record</th><th>Recuperi</th></tr></thead><tbody>{landing_rows}</tbody></table></div>
</div>
"""


def render_starship(starship):
    s = starship["metrics"]
    metric_html = "".join(
        [
            metric(s["eventi"], "eventi timeline"),
            metric(s["voli"], "voli integrati"),
            metric(s["catch"], "catch booster"),
            metric(s["reflight"], "reflight Super Heavy"),
        ]
    )
    flight_rows = "\n".join(
        f"<tr><td>{escape(str(row['flight']))}</td><td>{escape(str(row['data']))}</td><td>{escape(str(row['milestone']))}</td><td>{escape(str(row['esito']))}</td></tr>"
        for row in starship["flights"]
    )
    themes = "\n".join(
        f"""<div class="theme"><h3>{escape(str(row['tema']))}</h3><p>{escape(str(row['sintesi']))}</p><p class="muted">{escape(str(row['impatto']))}</p></div>"""
        for row in starship["themes"]
    )
    return f"""
<div class="metrics">{metric_html}</div>
<div class="dash-grid" style="margin-top:18px">
  <div class="panel"><h3>Voli integrati</h3><table><thead><tr><th>Flight</th><th>Data</th><th>Milestone</th><th>Esito</th></tr></thead><tbody>{flight_rows}</tbody></table></div>
  <div class="panel"><h3>Temi tecnici</h3><div class="theme-list">{themes}</div></div>
</div>
"""


def render_pad_card(pad):
    timeline = "\n".join(
        f"""<div class="pad-mini-event"><strong>{escape(str(event['date']))}</strong><span>{escape(str(event['text']))}</span></div>"""
        for event in pad["timeline"]
    )
    current = "\n".join(f"<li>{escape(str(line))}</li>" for line in pad["current"])
    return f"""
<article class="pad-card" id="{escape(pad['id'])}">
  <figure>
    <img src="{escape(pad['imageHref'])}" alt="Infografica {escape(pad['title'])}" loading="lazy">
  </figure>
  <div>
    <p class="badge">{escape(pad['region'])}</p>
    <h3>{escape(pad['title'])}</h3>
    <p>{escape(pad['subtitle'])}</p>
    <dl>
      <dt>Localita</dt><dd>{escape(pad['location'])}</dd>
      <dt>Nome</dt><dd>{escape(pad['full'])}</dd>
    </dl>
    <div class="actions" style="margin-top:14px">
      <button class="button secondary pad-open" type="button" data-pad="{escape(pad['id'])}">Apri sulla mappa</button>
    </div>
    <div class="pad-mini-timeline">{timeline}</div>
    <ul class="pad-current-list">{current}</ul>
  </div>
</article>
"""


def render_pad_page(data):
    pads = data["pads"]
    cards = "\n".join(render_pad_card(pad) for pad in pads)
    source_rows = "\n".join(f"<li>{escape(pad['title'])}: {escape(pad['source_note'])}</li>" for pad in pads)
    leaflet_head = '<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">'
    leaflet_script = f"""<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script>
const launchPads = {json_for_html(pads)};
let padMap;
let padMarkers = {{}};
let activeRegion = 'Tutti';

function padPopup(pad) {{
  const timeline = pad.timeline.slice(0, 4).map((event) => `<li><strong>${{event.date}}</strong>: ${{event.text}}</li>`).join('');
  const current = pad.current.slice(0, 4).map((line) => `<li>${{line}}</li>`).join('');
  return `
    <div class="pad-popup">
      <img src="${{pad.imageHref}}" alt="Infografica ${{pad.title}}">
      <div class="pad-popup-body">
        <span class="badge">${{pad.region}}</span>
        <h3>${{pad.title}}</h3>
        <p><strong>${{pad.full}}</strong><br>${{pad.location}}</p>
        <p>${{pad.subtitle}}</p>
        <ul>${{timeline}}</ul>
        <ul>${{current}}</ul>
      </div>
    </div>`;
}}

function padIcon(pad) {{
  const labels = {{
    'SLC-37B': '37B',
    'SLC-40': '40',
    'LC-39A': '39A',
    'SLC-4E': '4E',
    'SLC-6': '6',
    'BOCA CHICA PAD 1': 'P1',
    'BOCA CHICA PAD 2': 'P2'
  }};
  return L.divIcon({{
    className: 'pad-marker',
    html: `<div class="pad-marker-dot">${{labels[pad.title] || pad.title}}</div>`,
    iconSize: [34, 34],
    iconAnchor: [17, 17],
    popupAnchor: [0, -18]
  }});
}}

function renderPadSideList() {{
  const list = document.getElementById('pad-side-list');
  if (!list) return;
  list.innerHTML = '';
  launchPads
    .filter((pad) => activeRegion === 'Tutti' || pad.region === activeRegion)
    .forEach((pad) => {{
      const button = document.createElement('button');
      button.type = 'button';
      button.className = 'pad-side-item';
      button.dataset.pad = pad.id;
      button.innerHTML = `<strong>${{pad.title}}</strong><span>${{pad.location}}</span>`;
      button.addEventListener('click', () => openPad(pad.id));
      list.appendChild(button);
    }});
}}

function setPadActive(id) {{
  document.querySelectorAll('.pad-side-item').forEach((item) => item.classList.toggle('active', item.dataset.pad === id));
}}

function openPad(id) {{
  const pad = launchPads.find((item) => item.id === id);
  const marker = padMarkers[id];
  if (!pad || !marker) return;
  setPadActive(id);
  padMap.flyTo([pad.lat, pad.lng], pad.region === 'Florida' ? 9 : 10, {{duration: 0.9}});
  window.setTimeout(() => marker.openPopup(), 650);
}}

function filterPads(region) {{
  activeRegion = region;
  document.querySelectorAll('.pad-filter').forEach((button) => button.classList.toggle('active', button.dataset.region === region));
  launchPads.forEach((pad) => {{
    const marker = padMarkers[pad.id];
    if (!marker) return;
    const visible = region === 'Tutti' || pad.region === region;
    if (visible && !padMap.hasLayer(marker)) marker.addTo(padMap);
    if (!visible && padMap.hasLayer(marker)) marker.remove();
  }});
  renderPadSideList();
}}

function initPadMap() {{
  const mapEl = document.getElementById('pad-map');
  if (!mapEl || typeof L === 'undefined') return;
  padMap = L.map('pad-map', {{ zoomControl: true, minZoom: 3, maxZoom: 18 }}).setView([31.6, -96.4], 4);
  const osm = L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{
    attribution: '&copy; OpenStreetMap contributors'
  }}).addTo(padMap);
  const satellite = L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{{z}}/{{y}}/{{x}}', {{
    attribution: 'Tiles &copy; Esri'
  }});
  L.control.layers({{'OpenStreetMap': osm, 'Satellite': satellite}}, null, {{position: 'topright', collapsed: true}}).addTo(padMap);
  launchPads.forEach((pad) => {{
    const marker = L.marker([pad.lat, pad.lng], {{icon: padIcon(pad), title: pad.title}})
      .bindPopup(padPopup(pad), {{maxWidth: 390, minWidth: 290, className: 'custom-popup'}});
    marker.on('click', () => setPadActive(pad.id));
    marker.addTo(padMap);
    padMarkers[pad.id] = marker;
  }});
  document.querySelectorAll('.pad-filter').forEach((button) => {{
    button.addEventListener('click', () => filterPads(button.dataset.region));
  }});
  document.querySelectorAll('.pad-open').forEach((button) => {{
    button.addEventListener('click', () => {{
      document.getElementById('mappa-pad').scrollIntoView({{behavior: 'smooth', block: 'start'}});
      openPad(button.dataset.pad);
    }});
  }});
  renderPadSideList();
}}

window.addEventListener('load', initPadMap);
</script>"""
    body = f"""
{page_hero("Pad di lancio", "Infrastruttura SpaceX", "Mappa dei complessi di lancio SpaceX e schede operative ricavate dalle infografiche locali: storia del sito, fase attuale e ruolo dentro la cadenza futura.")}
<section id="mappa-pad">
  <div class="inner">
    <div class="section-head">
      <h2>Mappa operativa</h2>
      <p>Clicca un marker o una voce dell'elenco per aprire la scheda del pad sulla mappa. Sotto trovi anche l'archivio completo con le infografiche.</p>
    </div>
    <div class="pad-map-wrap">
      <aside class="pad-map-side">
        <div>
          <p class="badge">Pad SpaceX</p>
          <h3>Complessi attivi, in conversione o pianificati</h3>
          <p>Florida, California e Texas sono i tre poli che reggono Falcon, Dragon e Starship.</p>
        </div>
        <div class="pad-filter-row" aria-label="Filtri mappa pad">
          <button class="pad-filter active" type="button" data-region="Tutti">Tutti</button>
          <button class="pad-filter" type="button" data-region="Florida">Florida</button>
          <button class="pad-filter" type="button" data-region="California">California</button>
          <button class="pad-filter" type="button" data-region="Texas">Texas</button>
        </div>
        <div id="pad-side-list" class="pad-side-list"></div>
      </aside>
      <div id="pad-map" aria-label="Mappa dei pad di lancio SpaceX"></div>
    </div>
  </div>
</section>
<section>
  <div class="inner">
    <div class="section-head">
      <h2>Elenco pad</h2>
      <p>Le schede riprendono i contenuti delle infografiche nella cartella locale, in forma leggibile e navigabile.</p>
    </div>
    <div class="pad-list">{cards}</div>
  </div>
</section>
<section>
  <div class="inner">
    <div class="section-head">
      <h2>Fonti</h2>
      <p>Le schede si basano sulle fonti operative raccolte per le infografiche dei pad.</p>
    </div>
    <div class="source-panel">
      <p>Crediti immagini e fonti dettagliate sono conservati nel file locale <a href="../pad_di_lancio/fonti_pad_di_lancio.md">fonti_pad_di_lancio.md</a>.</p>
      <ul>{source_rows}</ul>
    </div>
  </div>
</section>
"""
    return shell("Pad di lancio", "pad-di-lancio", True, body, leaflet_script, leaflet_head)


def render_locations_page(data):
    locations = data["locations"]
    source_rows = "\n".join(f"<li>{escape(item['name'])}: {escape(item['source_note'])}</li>" for item in locations)
    leaflet_head = '<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">'
    leaflet_script = (
        '<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>\n'
        "<script>\n"
        "const spacexLocations = "
        + json_for_html(locations)
        + r""";
let locationMap;
let locationMarkers = {};
let locationPolygon;
let activeLocationFilter = 'Tutti';

function htmlEscape(value) {
  return String(value ?? '').replace(/[&<>"']/g, (char) => ({
    '&': '&amp;',
    '<': '&lt;',
    '>': '&gt;',
    '"': '&quot;',
    "'": '&#39;'
  }[char]));
}

function locationPopup(item) {
  const detailItems = (item.details || []).map((line) => `<li>${htmlEscape(line)}</li>`).join('');
  const classes = item.proposed ? 'location-popup proposed' : 'location-popup';
  return `
    <div class="${classes}">
      <div class="location-popup-head">
        <span class="badge">${htmlEscape(item.status)}</span>
        <h3>${htmlEscape(item.name)}</h3>
        <p>${htmlEscape(item.short)}</p>
      </div>
      <div class="location-popup-body">
        <div class="location-popup-box">
          <b>Ruolo</b>
          <p>${htmlEscape(item.role)}</p>
        </div>
        <div class="location-popup-box">
          <b>Indirizzo / area</b>
          <p>${htmlEscape(item.address)}</p>
        </div>
        <div class="location-popup-box">
          <b>Sintesi</b>
          <p>${htmlEscape(item.summary)}</p>
        </div>
        <div class="location-popup-box">
          <b>Dettagli</b>
          <ul>${detailItems}</ul>
        </div>
        <div class="location-popup-box">
          <b>Coordinate</b>
          <p>${htmlEscape(item.coords)}</p>
        </div>
      </div>
    </div>`;
}

function locationIcon(item) {
  const label = item.proposed ? 'LA' : 'SX';
  const cls = item.proposed ? 'location-dot proposed' : 'location-dot';
  return L.divIcon({
    className: 'pad-marker',
    html: `<div class="${cls}">${label}</div>`,
    iconSize: [34, 34],
    iconAnchor: [17, 17],
    popupAnchor: [0, -19]
  });
}

function renderLocationSideList() {
  const list = document.getElementById('location-side-list');
  if (!list) return;
  list.innerHTML = '';
  spacexLocations
    .filter((item) => activeLocationFilter === 'Tutti' || item.region === activeLocationFilter || (activeLocationFilter === 'Proposta' && item.proposed))
    .forEach((item) => {
      const button = document.createElement('button');
      button.type = 'button';
      button.className = `location-list-item ${item.proposed ? 'proposed' : ''}`;
      button.dataset.location = item.id;
      button.innerHTML = `<strong>${htmlEscape(item.name)}</strong><span>${htmlEscape(item.role)} · ${htmlEscape(item.short)}</span>`;
      button.addEventListener('click', () => openLocation(item.id));
      list.appendChild(button);
    });
}

function setLocationActive(id) {
  document.querySelectorAll('.location-list-item').forEach((item) => {
    item.classList.toggle('active', item.dataset.location === id);
  });
}

function openLocation(id) {
  const item = spacexLocations.find((candidate) => candidate.id === id);
  const marker = locationMarkers[id];
  if (!item || !marker) return;
  setLocationActive(id);
  const zoom = item.proposed ? 9 : 10;
  locationMap.flyTo([item.lat, item.lng], zoom, { duration: 0.9 });
  window.setTimeout(() => marker.openPopup(), 650);
}

function filterLocations(region) {
  activeLocationFilter = region;
  document.querySelectorAll('.pad-filter').forEach((button) => button.classList.toggle('active', button.dataset.region === region));
  spacexLocations.forEach((item) => {
    const marker = locationMarkers[item.id];
    if (!marker) return;
    const visible = region === 'Tutti' || item.region === region || (region === 'Proposta' && item.proposed);
    if (visible && !locationMap.hasLayer(marker)) marker.addTo(locationMap);
    if (!visible && locationMap.hasLayer(marker)) marker.remove();
  });
  if (locationPolygon) {
    const showPolygon = region === 'Tutti' || region === 'Louisiana' || region === 'Proposta';
    if (showPolygon && !locationMap.hasLayer(locationPolygon)) locationPolygon.addTo(locationMap);
    if (!showPolygon && locationMap.hasLayer(locationPolygon)) locationPolygon.remove();
  }
  renderLocationSideList();
}

function initLocationMap() {
  const mapEl = document.getElementById('location-map');
  if (!mapEl || typeof L === 'undefined') return;
  locationMap = L.map('location-map', { zoomControl: true, minZoom: 3, maxZoom: 18 }).setView([38.1, -98.4], 4);
  const osm = L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
    attribution: '&copy; OpenStreetMap contributors'
  }).addTo(locationMap);
  const satellite = L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', {
    attribution: 'Tiles &copy; Esri'
  });
  L.control.layers({ 'OpenStreetMap': osm, 'Satellite': satellite }, null, { position: 'topright', collapsed: true }).addTo(locationMap);

  spacexLocations.forEach((item) => {
    const marker = L.marker([item.lat, item.lng], { icon: locationIcon(item), title: item.name })
      .bindPopup(locationPopup(item), { maxWidth: 440, minWidth: 300, className: 'custom-popup' });
    marker.on('click', () => setLocationActive(item.id));
    marker.addTo(locationMap);
    locationMarkers[item.id] = marker;

    if (item.polygon) {
      locationPolygon = L.polygon(item.polygon, {
        color: '#f2b84b',
        weight: 2,
        dashArray: '7, 5',
        fillColor: '#f2b84b',
        fillOpacity: 0.12,
        opacity: 0.9
      }).bindPopup(locationPopup(item), { maxWidth: 440, minWidth: 300, className: 'custom-popup' });
      locationPolygon.addTo(locationMap);
    }
  });

  document.querySelectorAll('.pad-filter').forEach((button) => {
    button.addEventListener('click', () => filterLocations(button.dataset.region));
  });
  renderLocationSideList();
}

window.addEventListener('load', initLocationMap);
</script>"""
    )
    body = f"""
{page_hero("Localita SpaceX", "Mappa operativa", "Le principali localita SpaceX negli Stati Uniti: sedi, produzione, test, pad di lancio e l'ipotesi Louisiana non ancora attiva.")}
<section id="mappa-localita-spacex">
  <div class="inner">
    <div class="section-head">
      <h2>Mappa localita</h2>
      <p>La logica e la navigazione riprendono la mappa dei pad: clicca un marker o una voce dell'elenco per aprire un dettaglio leggibile. La Louisiana e segnata come proposta, non come struttura operativa.</p>
    </div>
    <div class="pad-map-wrap">
      <aside class="pad-map-side">
        <div>
          <p class="badge">SpaceX USA</p>
          <h3>Sedi, test, produzione e lancio</h3>
          <p>Il dataset finale contiene solo punti specificamente SpaceX.</p>
        </div>
        <div class="pad-filter-row" aria-label="Filtri mappa localita SpaceX">
          <button class="pad-filter active" type="button" data-region="Tutti">Tutti</button>
          <button class="pad-filter" type="button" data-region="California">California</button>
          <button class="pad-filter" type="button" data-region="Texas">Texas</button>
          <button class="pad-filter" type="button" data-region="Florida">Florida</button>
          <button class="pad-filter" type="button" data-region="Proposta">Proposta</button>
        </div>
        <div id="location-side-list" class="pad-side-list"></div>
      </aside>
      <div id="location-map" aria-label="Mappa delle localita SpaceX negli Stati Uniti"></div>
    </div>
  </div>
</section>
<section>
  <div class="inner">
    <div class="section-head">
      <h2>Note dati</h2>
      <p>Le coordinate sono pubbliche o rappresentative per aree estese. La localita Louisiana e deliberatamente trattata come scenario futuro/ipotetico.</p>
    </div>
    <div class="source-panel">
      <p>Dataset derivato dal file locale <strong>mappa_spacex_xai_usa.html</strong>, filtrando le sole localita SpaceX e mantenendo la proposta Louisiana.</p>
      <ul>{source_rows}</ul>
    </div>
  </div>
</section>
"""
    return shell("Localita SpaceX", "spacex", True, body, leaflet_script, leaflet_head)


def page_hero(title, eyebrow, copy, construction=False):
    badge = '<p class="badge">Pagina in costruzione</p>' if construction else ""
    return f"""
<section class="page-hero">
  <div class="inner">
    <p class="eyebrow">{escape(eyebrow)}</p>
    <h1>{escape(title)}</h1>
    <p class="subtitle">{escape(copy)}</p>
    {badge}
  </div>
</section>
"""


def render_spacex(data):
    f = data["falcon"]["metrics"]
    upcoming_exact = [item for item in data["upcoming"] if "net" not in (item.get("cat") or [])][:3]
    upcoming_block = (
        render_launch_cards(upcoming_exact, compact=True)
        if upcoming_exact
        else '<div class="panel"><p class="muted">Nessun T-0 puntuale disponibile al momento.</p></div>'
    )
    top_metrics = "".join(
        [
            metric(f"{f['lanci']:,}".replace(",", "."), "lanci Falcon principali"),
            metric(percent(f["tasso"]), "success rate storico"),
            metric(f["recuperiRiusciti"], "recuperi booster riusciti"),
            metric(data["starship"]["metrics"]["stato"], "stato Starship"),
        ]
    )
    body = f"""
{page_hero("SpaceX", "Sezione attiva", "L'area SpaceX raccoglie la parte viva del sito: agenda dei lanci, storico Falcon, riuso dei booster e sviluppo Starship.")}
<section>
  <div class="inner split">
    <article class="panel">
      <h2>Macchina operativa</h2>
      <p>SpaceX non viene trattata come una semplice azienda di lanci. Qui e il primo laboratorio della nuova industria spaziale: cadenza, costi, recupero, riuso, infrastrutture e sviluppo rapido nello stesso sistema.</p>
      <div class="actions spacex-actions">
        <a class="button history-link" href="storia-spacex.html">Storia</a>
        <a class="button" href="lanci-imminenti.html">Lanci imminenti</a>
        <a class="button secondary" href="storico-lanci.html">Storico lanci</a>
        <a class="button secondary" href="starship.html">Starship</a>
        <a class="button secondary" href="pad-di-lancio.html">Pad di lancio</a>
        <a class="button secondary" href="localita-spacex.html">Localita SpaceX</a>
      </div>
    </article>
    <aside class="panel"><h3>Cruscotto rapido</h3><div class="metrics">{top_metrics}</div></aside>
  </div>
</section>
<section>
  <div class="inner">
    <div class="section-head">
      <h2>Prossimi lanci</h2>
      <p>Anteprima dei T-0 piu vicini, ripulita dai voli gia avvenuti e allineata alla pagina agenda. Per finestre NET e fonti complete resta la sezione dedicata.</p>
    </div>
    {upcoming_block}
    <div class="actions">
      <a class="button" href="lanci-imminenti.html">Apri agenda completa</a>
      <a class="button secondary" href="storico-lanci.html">Vedi storico aggiornato</a>
    </div>
  </div>
</section>
"""
    return shell("SpaceX", "spacex", True, body, countdown_script())


def render_spacex_history_page(data):
    body = f"""
{page_hero("Storia SpaceX", "Sottosezione SpaceX", "Archivio narrativo della storia SpaceX, organizzato in parti selezionabili.")}
<section>
  <div class="inner">
    <div class="section-head">
      <h2>Parti disponibili</h2>
      <p>Il percorso storico resta separato dalla pagina operativa SpaceX. Ogni blocco apre un approfondimento autonomo.</p>
    </div>
    <div class="story-list">
      <a class="story-card" href="../documenti%20per%20sito/fondazione_spacex_fino_primo_falcon1.html">
        <img src="../documenti%20per%20sito/assets_fondazione/falcon1_flight4_launch.jpg" alt="Falcon 1 durante un lancio da Omelek Island" loading="lazy">
        <div>
          <small>Prima parte</small>
          <h3>Dalla fondazione fino al primo lancio</h3>
          <p>Dall'idea Mars Oasis alla nascita di SpaceX, dal nucleo tecnico iniziale al Falcon 1 e al primo volo da Omelek Island del 24 marzo 2006.</p>
          <span class="button secondary">Apri approfondimento</span>
        </div>
      </a>
      <a class="story-card" href="../documenti%20per%20sito/falcon1_dal_primo_fallimento_al_quarto_lancio.html">
        <img src="../documenti%20per%20sito/assets_fondazione/falcon1_flight4_launch.jpg" alt="Falcon 1 in volo durante il quarto lancio" loading="lazy">
        <div>
          <small>Seconda parte</small>
          <h3>Dal primo fallimento al quarto lancio</h3>
          <p>Dal Flight 1 perso su Omelek al successo orbitale del 28 settembre 2008: corrosione, slosh, spinta residua, crisi finanziaria e metodo di correzione SpaceX.</p>
          <span class="button secondary">Apri approfondimento</span>
        </div>
      </a>
      <a class="story-card" href="../documenti%20per%20sito/falcon1_successo_orbitale_contratto_crs.html">
        <img src="../documenti%20per%20sito/assets_fondazione/falcon1_flight4_launch.jpg" alt="Falcon 1 durante il lancio orbitale riuscito del 2008" loading="lazy">
        <div>
          <small>Terza parte</small>
          <h3>Dal primo successo orbitale al contratto CRS</h3>
          <p>Il quarto Falcon 1 diede a SpaceX l'orbita; il contratto NASA del dicembre 2008 diede all'azienda una traiettoria industriale. RazakSAT chiuse la storia del piccolo lanciatore.</p>
          <span class="button secondary">Apri approfondimento</span>
        </div>
      </a>
    </div>
  </div>
</section>
"""
    return shell("Storia SpaceX", "spacex", True, body)


def render_launches_page(data):
    launches = data["upcoming"]
    exact = [item for item in launches if "net" not in (item.get("cat") or [])]
    net = [item for item in launches if "net" in (item.get("cat") or [])]
    next_item = exact[0] if exact else (launches[0] if launches else None)
    if next_item:
        next_block = f"""
    <article class="next-launch">
      <div>
        <p class="badge">Prossimo in agenda</p>
        <h3>{escape(str(next_item.get('name') or 'Missione'))}</h3>
        <p>{escape(str(next_item.get('dateLabel') or 'Data da confermare'))}</p>
      </div>
      <div>
        <p><strong>{escape(str(next_item.get('rocket') or ''))}</strong></p>
        <p>{escape(str(next_item.get('site') or ''))}</p>
        <p>{escape(str(next_item.get('payload') or ''))}</p>
      </div>
    </article>"""
        next_label = str(next_item.get("dateLabel") or "Da confermare")
    else:
        next_block = ""
        next_label = "Da confermare"
    exact_section = (
        f"""
<section>
  <div class="inner">
    <div class="section-head">
      <h2>T-0 confermati</h2>
      <p>Missioni con orario puntuale gia indicato dalle fonti operative. Sono le card da ricontrollare piu spesso, perche finestre e meteo possono cambiare in poche ore.</p>
    </div>
    {render_launch_cards(exact)}
  </div>
</section>"""
        if exact
        else ""
    )
    net_section = (
        f"""
<section>
  <div class="inner">
    <div class="section-head">
      <h2>Finestre NET</h2>
      <p>Voli senza T-0 stabile: restano utili per orientarsi, ma non vanno letti come appuntamenti puntuali.</p>
    </div>
    {render_launch_cards(net, compact=True)}
  </div>
</section>"""
        if net
        else ""
    )
    body = f"""
{page_hero("Lanci imminenti", "Agenda SpaceX", "Agenda ripulita dai voli gia avvenuti: T-0 puntuali separati dalle finestre NET, con countdown solo dove ha senso.")}
<section>
  <div class="inner">
    <div class="section-head">
      <h2>Quadro rapido</h2>
      <p>Ultimo aggiornamento locale: {escape(str(data.get('generatedAt') or ''))}. Il volo Starlink Group 17-54 del 15 giugno 2026 e stato tolto dall'agenda perche gia avvenuto.</p>
    </div>
    <div class="agenda-strip">
      <div class="agenda-note"><b>{len(exact)}</b><span>lanci con T-0 puntuale ancora in agenda</span></div>
      <div class="agenda-note"><b>{len(net)}</b><span>finestre NET o missioni senza data stabile</span></div>
      <div class="agenda-note"><b>{escape(next_label)}</b><span>prossimo appuntamento mostrato</span></div>
    </div>
    {next_block}
  </div>
</section>
{exact_section}
{net_section}
"""
    return shell("Lanci imminenti", "lanci-imminenti", True, body, countdown_script())


def render_history_page(data):
    body = f"""
{page_hero("Storico lanci", "Dashboard Falcon", "Riepilogo operativo dello storico Falcon: cadenza, successi, pad, famiglie di lanciatore e recupero booster.")}
<section>
  <div class="inner">{render_falcon_dashboard(data['falcon'])}</div>
</section>
"""
    return shell("Storico lanci", "storico-lanci", True, body)


def render_starship_page(data):
    body = f"""
{page_hero("Starship", "Sviluppo SpaceX", "Produzione rapida, riuso completo, rifornimento orbitale e logistica lunare: Starship merita una pagina separata dalla porta generale SpaceX.")}
<section>
  <div class="inner">{render_starship(data['starship'])}</div>
</section>
"""
    return shell("Starship", "starship", True, body)


def render_placeholder(item):
    body = f"""
{page_hero(item['title'], item['status'], item['copy'], construction=True)}
<section>
  <div class="inner split">
    <article class="panel construction">
      <h2>In costruzione</h2>
      <p>Questa pagina e gia separata nella struttura del sito, ma il contenuto editoriale verra sviluppato in una fase successiva. Per ora resta come porta autonoma, con navigazione coerente e percorsi gia pronti per GitHub Pages.</p>
    </article>
    <aside class="panel">
      <h3>Prossimo sviluppo</h3>
      <p class="muted">Qui entreranno contesto, cronologia, dati, fonti e collegamenti alle pagine correlate.</p>
    </aside>
  </div>
</section>
"""
    return shell(item["title"], item["slug"], True, body)


def write(path, content):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def render():
    data = {
        "upcoming": extract_upcoming_launches(),
        "falcon": falcon_data(),
        "starship": starship_data(),
        "pads": pad_launch_data(),
        "locations": spacex_locations_data(),
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    write(CSS_DIR / "style.css", render_css())
    write(ROOT / "index.html", render_home())
    write(SECTIONS_DIR / "spacex.html", render_spacex(data))
    write(SECTIONS_DIR / "storia-spacex.html", render_spacex_history_page(data))
    write(SECTIONS_DIR / "lanci-imminenti.html", render_launches_page(data))
    write(SECTIONS_DIR / "storico-lanci.html", render_history_page(data))
    write(SECTIONS_DIR / "starship.html", render_starship_page(data))
    write(SECTIONS_DIR / "pad-di-lancio.html", render_pad_page(data))
    write(SECTIONS_DIR / "localita-spacex.html", render_locations_page(data))
    for item in PLACEHOLDER_SECTIONS:
        if item["slug"] != "spacex":
            write(SECTIONS_DIR / f"{item['slug']}.html", render_placeholder(item))


if __name__ == "__main__":
    render()
